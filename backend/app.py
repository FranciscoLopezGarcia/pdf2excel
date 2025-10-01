"""Flask entry point for the PDF to Excel service."""

import datetime
import io
import logging
import os
import time
import zipfile
from functools import wraps
from pathlib import Path
from typing import Dict, Iterable, List

import jwt
import pandas as pd
from flask import Flask, Response, jsonify, request, send_file
from flask_cors import CORS
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from config import CORS_ORIGINS, INPUT_DIR, MAX_UPLOAD_SIZE, OUTPUT_DIR, SECRET_KEY
from extractors.unificador import unir_consolidados
from extractors.universal_extractor import UniversalBankExtractor


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Directories and globals
# ---------------------------------------------------------------------------
INPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

PROGRESS_STATE: Dict[str, str] = {}


def _parse_origins(origins: str) -> Iterable[str] | str:
    items = [item.strip() for item in origins.split(",") if item.strip()]
    if not items or "*" in items:
        return "*"
    return items


# ---------------------------------------------------------------------------
# Flask application
# ---------------------------------------------------------------------------
app = Flask(__name__)
CORS(
    app,
    resources={r"/api/*": {"origins": _parse_origins(CORS_ORIGINS)}},
    supports_credentials=True,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def token_required(fn):
    """Decorator that validates a JWT bearer token."""

    @wraps(fn)
    def wrapper(*args, **kwargs):
        token = None

        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header.split(" ", 1)[1]

        if not token:
            token = request.args.get("token")

        if not token:
            logger.warning("Missing token in request")
            return jsonify({"error": "Token requerido"}), 401

        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            request.user = payload
            logger.info("Token valido para usuario %s", payload.get("username"))
        except Exception as exc:
            logger.error("Token invalido: %s", exc)
            return jsonify({"error": "Token invalido", "detail": str(exc)}), 401

        return fn(*args, **kwargs)

    return wrapper


def _update_progress(user: str, progress: int, status: str) -> None:
    PROGRESS_STATE[user] = f'{{"progress": {progress}, "status": "{status}"}}'


def _format_amounts(df: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    for column in columns:
        if column in df.columns:
            df[column] = df[column].apply(
                lambda value: (
                    f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    if pd.notna(value)
                    else "0,00"
                )
            )
    return df


def _highlight_observations(workbook_bytes: io.BytesIO) -> io.BytesIO:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        workbook = load_workbook(workbook_bytes)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if "observaciones" not in headers:
            return workbook_bytes

        obs_column = headers.index("observaciones") + 1
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=obs_column)
            if cell.value and str(cell.value).strip():
                for col_idx in range(1, sheet.max_column + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = yellow

        workbook_bytes = io.BytesIO()
        workbook.save(workbook_bytes)
        workbook_bytes.seek(0)
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Could not highlight observations: %s", exc)
        workbook_bytes.seek(0)
    return workbook_bytes


# ---------------------------------------------------------------------------
# Authentication
# ---------------------------------------------------------------------------
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(force=True)
    username = data.get("username")
    password = data.get("password")

    if username == "admin" and password == "admin123":
        role = "admin"
    elif username == "user" and password == "user123":
        role = "user"
    else:
        logger.warning("Credenciales invalidas para usuario %s", username)
        return jsonify({"error": "Credenciales invalidas"}), 401

    token = jwt.encode(
        {
            "username": username,
            "role": role,
            "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=8),
        },
        SECRET_KEY,
        algorithm="HS256",
    )

    return jsonify({"token": token, "role": role})


# ---------------------------------------------------------------------------
# Server Sent Events for progress feedback
# ---------------------------------------------------------------------------
@app.route("/api/progress")
@token_required
def progress():
    user = request.user["username"]

    def stream():
        last_payload = ""
        idle_ticks = 0
        timeout_ticks = 3000

        while idle_ticks < timeout_ticks:
            time.sleep(1)
            idle_ticks += 1

            current = PROGRESS_STATE.get(user)
            if current and current != last_payload:
                yield f"data: {current}\n\n"
                last_payload = current
                idle_ticks = 0

            if current and '"progress": 100' in current:
                break

        if idle_ticks >= timeout_ticks:
            yield 'data: {"progress": 0, "status": "Timeout - reconecta"}\n\n'

    response = Response(stream(), mimetype="text/event-stream")
    response.headers["Cache-Control"] = "no-cache"
    response.headers["Connection"] = "keep-alive"
    response.headers["X-Accel-Buffering"] = "no"
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Authorization"
    return response


# ---------------------------------------------------------------------------
# Conversion endpoint
# ---------------------------------------------------------------------------
@app.route("/api/convert", methods=["POST"])
@token_required
def convert_route():
    user = request.user.get("username")
    extractor = UniversalBankExtractor()
    uploaded_files: List[FileStorage] = request.files.getlist("files")

    if not uploaded_files:
        return jsonify({"error": "No se enviaron archivos"}), 400

    total_files = len(uploaded_files)
    _update_progress(user, 0, "Iniciando conversion")

    output = io.BytesIO()
    archive = zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED)

    combined_frames: List[pd.DataFrame] = []
    log_messages: List[str] = []

    for index, storage in enumerate(uploaded_files, start=1):
        filename = secure_filename(storage.filename or f"archivo_{index}.pdf")
        temp_path = INPUT_DIR / filename
        storage.save(temp_path)

        try:
            if temp_path.stat().st_size > MAX_UPLOAD_SIZE:
                raise ValueError("Archivo supera el tamano permitido")

            _update_progress(user, int(((index - 1) / total_files) * 100), f"Procesando {filename}")

            dataframe = extractor.extract_from_pdf(str(temp_path), filename_hint=filename)

            if dataframe is None or dataframe.empty:
                message = f"{filename}: ERROR - sin transacciones"
                archive.writestr(f"{filename}-ERROR.txt", message)
                log_messages.append(message)
                continue

            export_frame = dataframe.copy()
            export_frame = _format_amounts(export_frame, ["debitos", "creditos", "saldo"])

            excel_buffer = io.BytesIO()
            export_frame.to_excel(excel_buffer, index=False, sheet_name="Transactions")
            excel_buffer.seek(0)
            excel_buffer = _highlight_observations(excel_buffer)

            archive.writestr(f"{Path(filename).stem}.xlsx", excel_buffer.getvalue())

            dataframe["archivo"] = filename
            combined_frames.append(dataframe)

            observations = dataframe.get("observaciones", pd.Series(dtype=str))
            problems = int(observations.apply(lambda value: bool(str(value).strip())).sum())
            message = f"{filename}: OK - {len(dataframe)} filas ({problems} con observaciones)"
            log_messages.append(message)
        except Exception as exc:
            message = f"{filename}: ERROR - {exc}"
            archive.writestr(f"{filename}-ERROR.txt", message)
            log_messages.append(message)
            logger.error("Error procesando %s: %s", filename, exc, exc_info=True)
        finally:
            if temp_path.exists():
                temp_path.unlink()

        _update_progress(user, int((index / total_files) * 100), f"Completado {filename}")

    if combined_frames:
        combined_df = pd.concat(combined_frames, ignore_index=True)
        combined_buffer = io.BytesIO()
        combined_df.to_excel(combined_buffer, index=False, sheet_name="Consolidado")
        archive.writestr("consolidado.xlsx", combined_buffer.getvalue())
        log_messages.append("Consolidado generado con exito")
    else:
        log_messages.append("No se pudo generar consolidado")

    archive.writestr("log.txt", "\n".join(log_messages))
    archive.close()
    output.seek(0)

    _update_progress(user, 100, "Finalizado")

    return send_file(
        output,
        mimetype="application/zip",
        as_attachment=True,
        download_name="resultado.zip",
    )


# ---------------------------------------------------------------------------
# Logs (demo)
# ---------------------------------------------------------------------------
@app.route("/api/logs", methods=["GET"])
@token_required
def logs():
    data = [
        {"user": "fran", "date": "2025-09-24", "ok": 3, "errors": 1, "reason": "OCR fallo"},
        {"user": "ana", "date": "2025-09-23", "ok": 5, "errors": 0, "reason": ""},
    ]
    return jsonify(data)


# ---------------------------------------------------------------------------
# Merge endpoint
# ---------------------------------------------------------------------------
@app.route("/api/unificar", methods=["POST"])
@token_required
def unificar():
    uploaded_files: List[FileStorage] = request.files.getlist("files")
    if not uploaded_files:
        return jsonify({"error": "No se enviaron archivos"}), 400

    temp_files: List[Path] = []
    try:
        for storage in uploaded_files:
            filename = secure_filename(storage.filename or "archivo.xlsx")
            temp_path = INPUT_DIR / filename
            storage.save(temp_path)
            temp_files.append(temp_path)

        output_path = OUTPUT_DIR / "consolidado_anual.xlsx"
        unir_consolidados([str(path) for path in temp_files], str(output_path))

        return send_file(
            output_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="consolidado_anual.xlsx",
        )
    finally:
        for path in temp_files:
            if path.exists():
                path.unlink()


if __name__ == "__main__":
    logger.info("Iniciando servidor Flask")
    app.run(host="0.0.0.0", port=5000, debug=True, threaded=True)
