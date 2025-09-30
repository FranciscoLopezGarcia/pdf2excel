#!/usr/bin/env python3
"""
Simple PDF to Excel Bank Statement Extractor
Drop PDFs in input folder, get Excel files in output folder.
No configuration needed.
"""

import os
import sys
from pathlib import Path
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from .universal_extractor import UniversalBankExtractor

# Setup simple logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('pdf2xls.log')
    ]
)

log = logging.getLogger("pdf2xls")

def highlight_problem_rows(excel_file: str):
    """
    Highlight rows with observaciones in Excel output.
    Rows with problems get a yellow background.
    """
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Find observaciones column (should be last column)
        headers = [cell.value for cell in ws[1]]
        try:
            obs_col_idx = headers.index('observaciones') + 1  # 1-indexed for openpyxl
        except ValueError:
            log.warning(f"No 'observaciones' column found in {excel_file}")
            return
        
        # Yellow fill for problem rows
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Iterate through data rows (skip header)
        for row_idx in range(2, ws.max_row + 1):
            obs_cell = ws.cell(row=row_idx, column=obs_col_idx)
            if obs_cell.value and str(obs_cell.value).strip():
                # Highlight entire row
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
        
        wb.save(excel_file)
        log.info(f"Highlighted problem rows in {excel_file}")
        
    except Exception as e:
        log.warning(f"Could not highlight problem rows in {excel_file}: {e}")

def format_excel_output(df: pd.DataFrame, output_file: str):
    """
    Save DataFrame to Excel with proper formatting:
    - Argentine number format for amounts
    - Highlight problem rows
    """
    # Format amounts to Argentine format (1.234,56)
    for col in ['debitos', 'creditos', 'saldo']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if pd.notna(x) else '0,00')
    
    # Save to Excel
    df.to_excel(output_file, index=False, sheet_name="Transactions")
    
    # Apply highlighting
    highlight_problem_rows(output_file)

def main():
    # Setup paths - input is outside extractors folder
    script_dir = Path(__file__).parent
    project_root = script_dir.parent  # Go up one level from extractors
    input_dir = project_root / "input"
    output_dir = project_root / "output"
    
    # Create directories
    input_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)
    
    log.info(f"Input directory: {input_dir}")
    log.info(f"Output directory: {output_dir}")
    
    # Find PDF files recursively in all subdirectories
    pdf_files = list(input_dir.rglob("*.pdf"))
    if not pdf_files:
        log.warning(f"No PDF files found in {input_dir} (searched recursively)")
        print(f"\nNo PDF files found in {input_dir} or subdirectories")
        print("Please add PDF files to the input folder and run again.")
        return
    
    log.info(f"Found {len(pdf_files)} PDF files to process")
    
    # Initialize extractor
    extractor = UniversalBankExtractor()
    
    # Process each PDF
    processed = 0
    failed = 0
    
    for pdf_file in pdf_files:
        try:
            # Create relative path for output structure
            rel_path = pdf_file.relative_to(input_dir)
            output_subdir = output_dir / rel_path.parent
            output_subdir.mkdir(parents=True, exist_ok=True)
            
            log.info(f"Processing: {rel_path}")
            print(f"\nProcessing: {rel_path}")
            
            # Extract data with filename hint for year inference
            df = extractor.extract_from_pdf(str(pdf_file), filename_hint=str(pdf_file))
            
            if df.empty:
                log.warning(f"No data extracted from {rel_path}")
                print(f"  ⚠️  No transactions found")
                failed += 1
                continue
            
            # Generate output filename (preserve folder structure)
            output_file = output_subdir / f"{pdf_file.stem}.xlsx"
            
            # Save to Excel with formatting
            format_excel_output(df, str(output_file))
            
            # Count problems
            problem_count = df['observaciones'].apply(lambda x: bool(str(x).strip())).sum()
            
            log.info(f"Saved {len(df)} transactions to {output_file.relative_to(output_dir)}")
            print(f"  ✅ Extracted {len(df)} transactions → {output_file.relative_to(output_dir)}")
            
            if problem_count > 0:
                print(f"  ⚠️  {problem_count} rows flagged with problems (highlighted in yellow)")
            
            processed += 1
            
        except Exception as e:
            log.error(f"Failed to process {pdf_file.relative_to(input_dir)}: {e}", exc_info=True)
            print(f"  ❌ Failed: {e}")
            failed += 1
    
    # Summary
    log.info(f"Processing complete: {processed} successful, {failed} failed")
    print(f"\n" + "="*50)
    print(f"SUMMARY:")
    print(f"  Processed: {processed} files")
    print(f"  Failed: {failed} files")
    print(f"  Output folder: {output_dir}")
    
    if processed > 0:
        print(f"\n✅ Check the output folder for your Excel files!")
        print(f"   Rows with problems are highlighted in yellow.")
    
    # Wait for user input before closing (useful when double-clicking)
    if len(sys.argv) == 1:  # No command line arguments = probably double-clicked
        input("\nPress Enter to exit...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nCancelled by user")
    except Exception as e:
        log.error(f"Unexpected error: {e}", exc_info=True)
        print(f"\n❌ Unexpected error: {e}")
        input("Press Enter to exit...")